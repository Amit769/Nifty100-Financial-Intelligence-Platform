"""
Peer Comparison Excel Report
============================

Generates peer_comparison.xlsx with:
- One sheet per peer group
- Financial metrics
- Percentile ranks
- Benchmark highlighting
- Peer median summary
"""

from pathlib import Path

import sqlite3
import pandas as pd

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


PROJECT_ROOT = Path(__file__).resolve().parents[2]

DATABASE_PATH = (
    PROJECT_ROOT
    / "database"
    / "nifty100.db"
)

OUTPUT_PATH = (
    PROJECT_ROOT
    / "output"
    / "peer_comparison.xlsx"
)


PEER_METRICS = {
    "ROE": "return_on_equity_pct",
    "ROCE": "return_on_capital_employed_pct",
    "NPM": "net_profit_margin_pct",
    "D/E": "debt_to_equity",
    "FCF": "free_cash_flow_cr",
    "PAT CAGR 5yr": "pat_cagr_5yr",
    "Revenue CAGR 5yr": "revenue_cagr_5yr",
    "EPS CAGR 5yr": "eps_cagr_5yr",
    "Interest Coverage": "interest_coverage",
    "Asset Turnover": "asset_turnover",
}

def load_peer_report_data():
    """
    Load peer groups, company names,
    financial metrics, and percentile ranks.
    """

    connection = sqlite3.connect(
        DATABASE_PATH
    )

    try:

        peers = pd.read_sql_query(
            """
            SELECT
                company_id,
                peer_group_name,
                is_benchmark
            FROM peer_groups
            """,
            connection,
        )

        companies = pd.read_sql_query(
            """
            SELECT
                id AS company_id,
                company_name
            FROM companies
            """,
            connection,
        )

        ratios = pd.read_sql_query(
            """
            SELECT *
            FROM financial_ratios
            """,
            connection,
        )

        percentiles = pd.read_sql_query(
            """
            SELECT
                company_id,
                peer_group_name,
                metric,
                percentile_rank
            FROM peer_percentiles
            """,
            connection,
        )

    finally:

        connection.close()

    ratios["year_num"] = (
        ratios["year"]
        .astype(str)
        .str.extract(
            r"(\d{4})"
        )[0]
        .astype(float)
    )

    ratios = (
        ratios
        .sort_values(
            [
                "company_id",
                "year_num",
            ]
        )
        .drop_duplicates(
            "company_id",
            keep="last",
        )
    )

    data = peers.merge(
        companies,
        on="company_id",
        how="left",
    )

    data = data.merge(
        ratios[
            [
                "company_id",
                *PEER_METRICS.values(),
            ]
        ],
        on="company_id",
        how="left",
    )

    return data, percentiles 

def generate_peer_comparison_report():
    """
    Generate peer_comparison.xlsx with one sheet
    per peer group.
    """

    data, percentiles = (
        load_peer_report_data()
    )

    output_frames = {}

    for peer_group_name, group in data.groupby(
        "peer_group_name"
    ):

        report = group[
            [
                "company_id",
                "company_name",
                "is_benchmark",
                *PEER_METRICS.values(),
            ]
        ].copy()

        report = report.rename(
            columns={
                value: key
                for key, value
                in PEER_METRICS.items()
            }
        )

        group_percentiles = percentiles[
            percentiles[
                "peer_group_name"
            ]
            == peer_group_name
        ]

        percentile_wide = (
            group_percentiles
            .pivot(
                index="company_id",
                columns="metric",
                values="percentile_rank",
            )
            .reset_index()
        )

        percentile_wide = (
            percentile_wide.rename(
                columns={
                    metric: f"{metric}_Percentile"
                    for metric in PEER_METRICS.keys()
                }
            )
        )

        report = report.merge(
            percentile_wide,
            on="company_id",
            how="left",
        )

        report = report.sort_values(
            "company_id"
        )

        output_frames[
            peer_group_name
        ] = report

    OUTPUT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:

        for sheet_name, report in (
            output_frames.items()
        ):

            safe_sheet_name = (
                sheet_name[:31]
            )

            report.to_excel(
                writer,
                sheet_name=safe_sheet_name,
                index=False,
            )

    return OUTPUT_PATH

def format_peer_comparison_report():
    """
    Apply percentile colour coding,
    benchmark highlighting, and median summary rows.
    """

    workbook = load_workbook(
        OUTPUT_PATH
    )

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    yellow_fill = PatternFill(
        fill_type="solid",
        fgColor="FFEB9C",
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    benchmark_fill = PatternFill(
        fill_type="solid",
        fgColor="FFD966",
    )

    for worksheet in workbook.worksheets:

        headers = [
            cell.value
            for cell in worksheet[1]
        ]

        percentile_columns = [
            index + 1
            for index, header in enumerate(
                headers
            )
            if header
            and str(header).endswith(
                "_Percentile"
            )
        ]

        benchmark_column = (
            headers.index(
                "is_benchmark"
            )
            + 1
        )

        max_row = worksheet.max_row

        # Add median summary row
        summary_row = max_row + 2

        worksheet.cell(
            row=summary_row,
            column=1,
            value="PEER GROUP MEDIAN",
        )

        for column in range(
            4,
            worksheet.max_column + 1,
        ):

            values = []

            for row in range(
                2,
                max_row + 1,
            ):

                value = worksheet.cell(
                    row=row,
                    column=column,
                ).value

                if isinstance(
                    value,
                    (
                        int,
                        float,
                    ),
                ):

                    values.append(
                        value
                    )

            if values:

                worksheet.cell(
                    row=summary_row,
                    column=column,
                    value=pd.Series(
                        values
                    ).median(),
                )

        # Colour percentile cells
        for row in range(
            2,
            max_row + 1,
        ):

            # Benchmark row
            benchmark_value = worksheet.cell(
                row=row,
                column=benchmark_column,
            ).value

            if benchmark_value == 1:

                for column in range(
                    1,
                    worksheet.max_column + 1,
                ):

                    worksheet.cell(
                        row=row,
                        column=column,
                    ).fill = benchmark_fill

            # Percentile colours
            for column in percentile_columns:

                cell = worksheet.cell(
                    row=row,
                    column=column,
                )

                value = cell.value

                if value is None:
                    continue

                if value >= 75:
                    cell.fill = green_fill

                elif value <= 25:
                    cell.fill = red_fill

                else:
                    cell.fill = yellow_fill

        # Header formatting
        for cell in worksheet[1]:

            cell.font = cell.font.copy(
                bold=True
            )

        # Summary row formatting
        for cell in worksheet[
            summary_row
        ]:

            cell.font = cell.font.copy(
                bold=True
            )

        # Column widths
        for column_cells in worksheet.columns:

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            worksheet.column_dimensions[
                column_letter
            ].width = 18

    workbook.save(
        OUTPUT_PATH
    )

    return OUTPUT_PATH