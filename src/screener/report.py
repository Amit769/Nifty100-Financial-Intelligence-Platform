"""
Screener Excel Report
=====================

Generates screener_output.xlsx with one sheet
for each preset screener.
"""

from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from src.screener.engine import (
    load_screener_presets,
    run_preset_screener,
)
from src.screener.result_formatter import (
    format_screener_results,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]

OUTPUT_PATH = (
    PROJECT_ROOT
    / "output"
    / "screener_output.xlsx"
)


def generate_screener_report():
    """
    Generate screener_output.xlsx with one
    sheet per preset screener.
    """

    OUTPUT_PATH.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    presets = load_screener_presets()

    output_frames = {}

    for preset_name in presets:

        result = run_preset_screener(
            preset_name
        )

        report = format_screener_results(
            result
        )

        output_frames[preset_name] = report

    with pd.ExcelWriter(
        OUTPUT_PATH,
        engine="openpyxl",
    ) as writer:

        for sheet_name, report in output_frames.items():

            safe_sheet_name = (
                sheet_name[:31]
            )

            report.to_excel(
                writer,
                sheet_name=safe_sheet_name,
                index=False,
            )

    format_screener_report()

    return OUTPUT_PATH


def format_screener_report():
    """
    Apply basic formatting to the screener workbook.
    """

    workbook = load_workbook(
        OUTPUT_PATH
    )

    green_fill = PatternFill(
        fill_type="solid",
        fgColor="C6EFCE",
    )

    red_fill = PatternFill(
        fill_type="solid",
        fgColor="FFC7CE",
    )

    for worksheet in workbook.worksheets:

        worksheet.freeze_panes = "A2"

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = (
                get_column_letter(
                    column_cells[0].column
                )
            )

            for cell in column_cells:

                if cell.value is not None:

                    max_length = max(
                        max_length,
                        len(str(cell.value)),
                    )

            worksheet.column_dimensions[
                column_letter
            ].width = min(
                max_length + 2,
                30,
            )

        headers = {
            cell.value: cell.column
            for cell in worksheet[1]
        }

        score_column = headers.get(
            "composite_quality_score"
        )

        if score_column:

            for row in range(
                2,
                worksheet.max_row + 1,
            ):

                cell = worksheet.cell(
                    row=row,
                    column=score_column,
                )

                if (
                    cell.value is not None
                    and cell.value >= 50
                ):

                    cell.fill = green_fill

                else:

                    cell.fill = red_fill

    workbook.save(
        OUTPUT_PATH
    )


if __name__ == "__main__":

    generate_screener_report()